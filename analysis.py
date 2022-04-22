"""
Application : Kaggle Data Analysis 
File name   : analysis.py
Authors     : Jacob Summerville
Description : This file performs a linear regression and correlation matrix 
              analysis with the provided dataset
"""

import argparse
import datetime
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import os
import pandas as pd
from scipy import stats
import seaborn as sn
from sklearn import datasets, linear_model
from sklearn.linear_model import LinearRegression
import statsmodels.api as sm
import time

# Constants for analysis
PVALUE_CUTOFF = 0.05
CORRELATION_CUTOFF = 0.5
HIGH_CORRELATION_CUTOFF = 0.7

OUTPUT_DIR = 'output'
summary_filename = os.path.join(OUTPUT_DIR, 'summary.csv')

DEPEND_VAR = 'SalePrice'

def getArguments():
    """ Read in command line arguments """
    parser = argparse.ArgumentParser(description='This program performs an analysis')

    parser.add_argument('-f', '--filepath', help='Filepath of the excel file')
    parser.add_argument('-s', '--sheet', help='Sheet name for the desired data')
    parser.add_argument('-a', '--all', help='Run analysis for all sheets', action="store_true")

    return parser.parse_args()


def create_summary_file():
    """ Create the file to save summary data to """
    if not os.path.isdir(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    summary_file = open(summary_filename, 'w')
    summary_file.write('Iteration, Metrics, P-Value (<), Correlation (>), Multicollinearity (>), R-Squared, Average Error, MAPE\n')
    summary_file.close()


def sort_pvalue(val):
    """ Sort p-values by value """
    return val[1] 


class Analysis:
    def __init__(self, filename, sheet, verbose = None):
        self.raw_data_file = filename
        self.sheet_name = sheet
        self.verbose = verbose
        self.file_dir = os.path.join(OUTPUT_DIR, self.sheet_name, '')

        self.retained_metrics = {
            'pvalue': {},
            'correlation': {}
        }

        start_time = time.time()

        self.create_output_file()
        
        if (self.read_file_data() == 1):
            return
        
        self.linear_regression()
        self.test_model()
        self.pvalues()
        self.correlated()
        self.correlation_matrix()
        self.correlation_chart()
        self.keep_similar()
        self.close_output_file()

        completion_time = '%.2f' % float(time.time() - start_time)
        print(sheet + ' Analysis Complete (' + completion_time + ' seconds)')


    def create_output_file(self):
        """ Create the file to save output data to """
        if not os.path.isdir(self.file_dir):
            os.makedirs(self.file_dir)

        # name the file using the current date and time
        #today = datetime.datetime.now()
        #self.current_time = today.strftime("%Y%m%d_%H%M%S")
        filename = self.file_dir + self.sheet_name + '.txt'
        self.output_file = open(filename, 'w')


    def write_output(self, text):
        """ Write text to the output file """
        self.output_file.write(text + '\n')


    def close_output_file(self):
        """ Save and close output file """
        self.output_file.close()


    def read_file_data(self):
        """ Read the excel file data """
        try:
            df = pd.read_excel(self.raw_data_file, sheet_name=self.sheet_name)
            self.data_columns = df.columns
            half_point = (len(df[self.data_columns[0]]) // 2)
        except:
            print('Error reading ' + self.sheet_name)
            return 1

        self.train_df = df.iloc[:half_point]
        self.test_df = df.iloc[half_point:]

        return 0


    def linear_regression(self):
        """ Perform linear regression and output the summary """
        X = self.train_df[self.data_columns[1:]]
        Y = self.train_df[DEPEND_VAR]

        X = sm.add_constant(X)
        self.regression_model = sm.OLS(Y, X).fit()

        self.regression_coeff = self.regression_model.params

        if self.verbose:
            self.write_output(str(self.regression_model.summary()))
            self.write_output('\n' + ('=' * 40))

        # generate the correlation matrix dataframe
        self.correlation_df = self.train_df.corr()


    def test_model(self):
        """ Use the coefficients to test the remaining test data """
        data_points = 0
        avg_error = 0
        mape = 0

        for index, row in self.test_df.iterrows():
            estimated_sale_price = self.regression_coeff[0]

            for i in range(1, len(self.test_df.columns)):
                estimated_sale_price += row[i] * self.regression_coeff[i]

            absolute_error = abs(row[0] - estimated_sale_price)
            ape = absolute_error / row[0] 

            data_points += 1
            avg_error += absolute_error
            mape += ape

        r_squared_str = '{:3f}'.format(self.regression_model.rsquared)
        avg_error_str_clean = '${:,.2f}'.format(avg_error / data_points)
        avg_error_str = '${:.2f}'.format(avg_error / data_points)
        mape_str = '{:.3f}%'.format((mape / data_points) * 100)

        self.write_output('\nR-Squared: ' + r_squared_str)
        self.write_output('Average Error: ' + avg_error_str_clean)
        self.write_output('MAPE: ' + mape_str)

        # write summary
        summary_file = open(summary_filename, 'a')
        summary_file.write(self.sheet_name + ',' + str(len(self.test_df.columns) - 1) + \
            ',' + str(PVALUE_CUTOFF) + ',' + str(CORRELATION_CUTOFF) + ',' + \
            str(HIGH_CORRELATION_CUTOFF) + ',' + r_squared_str + ',' + \
            avg_error_str + ',' + mape_str + '\n')
        summary_file.close()


    def pvalues(self):
        """ Find p-values and display the ones above and below the cutoff value """
        pvalues_meeting_req = []

        # get all p-values below the specified cutoff
        for i in range(1, len(self.regression_model.pvalues)):
            if self.regression_model.pvalues[i] < PVALUE_CUTOFF:
                self.retained_metrics['pvalue'][self.data_columns[i]] = self.regression_model.pvalues[i]

        self.write_output('\nKeep based on regression: (Cutoff of ' + str(PVALUE_CUTOFF) + ')')
        self.write_output('------------------------------------------')

        if len(self.retained_metrics['pvalue']) == 0:
            self.write_output('None')
            return

        # sort by pvalue
        self.retained_metrics['pvalue'] = dict(sorted(self.retained_metrics['pvalue'].items(), 
                                                key=lambda item: item[1]))

        for metric, pvalue in self.retained_metrics['pvalue'].items():
            self.write_output('{:15s}: {:.2e}'.format(metric, pvalue))


    def correlation_matrix(self):
        """ Create a correlation matrix and display it """
        sn.heatmap(self.correlation_df, vmin=0.5, vmax=1, annot=True, cmap='BrBG')

        figure = plt.gcf()
        figure.set_size_inches(30, 20)
        plt.title("Correlation Matrix")
        plt.savefig(self.file_dir + self.sheet_name + '-Matrix.pdf')
        
        # clear plot to prevent future interference
        plt.clf()


    def correlation_chart(self):
        """ Create a correlation matrix with the dependent variable """
        correlation_values = self.correlation_df[[DEPEND_VAR]].sort_values(by=DEPEND_VAR, ascending=False)

        heatmap = sn.heatmap(correlation_values, vmin=-1, vmax=1, annot=True, cmap='BrBG')
        heatmap.set_title('Features Correlating with ' + DEPEND_VAR, pad=16);

        figure = plt.gcf()
        figure.set_size_inches(15, 12)
        plt.savefig(self.file_dir + self.sheet_name + '-Correlation.pdf')


    def correlated(self):
        """ Analyze correlated values and parse the ones to keep """
        sorted_df = self.correlation_df.sort_values(by=DEPEND_VAR, ascending=False)

        correlated_columns = sorted_df.index[sorted_df[DEPEND_VAR] > CORRELATION_CUTOFF].tolist()
        self.retained_metrics['correlation'] = dict.fromkeys(correlated_columns , None)

        self.write_output('\nKeep based on correlation: (Cutoff of ' + str(CORRELATION_CUTOFF) + ')')
        self.write_output('------------------------------------------')

        if len(self.retained_metrics['correlation']) == 0:
            self.write_output('None')
            return

        for metric in correlated_columns:
            if metric == DEPEND_VAR:
                continue

            correlation_value = sorted_df[DEPEND_VAR].loc[[metric]].item()
            formatted_output = '{:15s}: {:.2f}'.format(metric, correlation_value)

            # find multicollinearity between the primary metrics
            secondary_correlations = sorted_df.index[sorted_df[metric] > HIGH_CORRELATION_CUTOFF].tolist()
            mc_metrics = []

            for mc_metric in secondary_correlations: 
                if mc_metric in self.retained_metrics['correlation'] and mc_metric != DEPEND_VAR and mc_metric != metric:
                    mc_metrics.append(mc_metric)

            if len(mc_metrics) > 0:
                self.write_output(formatted_output + ' (Multicollinearity: ' + ', '.join(mc_metrics) + ')')
            else:
                self.write_output(formatted_output)


    def keep_similar(self):
        """ List off p-value and correlation values that should both be kept """
        self.write_output('\nKeep both:')
        self.write_output('--------------')

        match_found = False

        for metric in self.retained_metrics['correlation']:
            if metric in self.retained_metrics['pvalue']:
                self.write_output(metric)
                match_found = True

        if not match_found:
            self.write_output('None')
    

def main():
    """ Main function for the analysis """
    args = getArguments()

    filepath = 'train.xlsx'
    sheets = []

    if args.filepath and os.path.isfile(args.filepath): 
       filepath = args.filepath 
    
    if args.sheet:   
        sheets = [args.sheet]
    elif args.all:
        wb = load_workbook(filepath, read_only=True, keep_links=False)
        sheets = wb.sheetnames
    else:
        print('Please provide the sheet name for the data (Example: -s IT1)')
        return

    create_summary_file()

    for sheet in sheets:
        Analysis(filepath, sheet, verbose=True)

if __name__ == '__main__':
    main()